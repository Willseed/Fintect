# _*_ coding: UTF-8 _*_

import time
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


def init():
    year_list = [i for i in range(92, 107)]
    twse_list = ['2885'] #,'2885','2886','2887'
    tpex_list = ['4102','4103','4104','4105','4106','4107','4108','4109','4110','4111','4113','4114','4205']
    #print(len(year_list))
    #print(len(tpex_list))
    return year_list, twse_list, tpex_list

playwright_runtime = None
browser = None
page = None
popup = None

def current_page():
    return popup if popup is not None else page

def driver_open():
    global playwright_runtime, browser, page, popup
    playwright_runtime = sync_playwright().start()
    browser = playwright_runtime.chromium.launch(headless=False)
    page = browser.new_page()
    popup = None
    page.goto('https://mopsov.twse.com.tw/mops/web/t05st01')
    return page

def driver_close(browser_arg=None):
    global playwright_runtime, browser, page, popup
    if browser is not None:
        browser.close()
    if playwright_runtime is not None:
        playwright_runtime.stop()
    playwright_runtime = None
    browser = None
    page = None
    popup = None

def input_text(index, xpath):
    inputbox = page.locator('xpath=' + xpath)
    inputbox.fill(str(index))
    inputbox.press('Enter')

def WebWaitXpath(s):
    try:
        current_page().locator('xpath=' + s).first.wait_for(state='visible', timeout=2000)
        return True
    except Exception as e:
        print(e)

def ChangeToPopUpWindow(index):
    global popup
    with page.expect_popup() as popup_info:
        page.locator('xpath=//*[@id="t05st01_fm"]/table/tbody/tr[' + str(index) + ']/td[6]/input').click()
    popup = popup_info.value

def BackToSourceWindow(window_before):
    global popup
    if popup is not None:
        popup.close()
        popup = None

def ListToDict(length, l_title, l_content):
    d_details = {}
    for i in range(0, length):
        d_details[l_title[i]] = l_content[i]
    return d_details

def get_data():
    try:
        time.sleep(10) #等待5s 再次點擊
        WebWaitXpath('//*[@id="table01"]//table[contains(@class,"hasBorder")]') #等待元件讀取
        l_title = []
        l_content = []
        table_path = '//*[@id="table01"]//table[contains(@class,"hasBorder")]/tbody'
        target = current_page()
        col_path = target.locator('xpath=' + table_path + '/tr') #欄位置
        for i in range(1, col_path.count() + 1):
            row_path = target.locator('xpath=' + table_path + '/tr[' + str(i) + ']/td') #列位置
            for j in range(1, row_path.count() + 1):
                cell = target.locator('xpath=' + table_path + '/tr[' + str(i) + ']/td[' + str(j) + ']')
                if not (j % 2 == 0):
                    title = cell.inner_text() #標題
                    l_title.append(title)
                elif(i == col_path.count() and j == row_path.count()):
                    content = cell.inner_text().split('\n') #說明部分切割成List
                    for k in content:
                        k.lstrip().rstrip()
                    l_content.append(content)
                else:
                    content = cell.inner_text().lstrip().rstrip() #內容 去左右空白
                    l_content.append(content)
        return ListToDict(len(l_title), l_title, l_content)
    except:
        print('Get Data Error!')
        return False

def input_data(sheet, d_details):
    data = [['序號', d_details['序號']],
        ['發言日期', d_details['發言日期']],
        ['發言時間', d_details['發言時間']],
        ['發言人', d_details['發言人']],
        ['發言人職稱', d_details['發言人職稱']],
        ['發言人電話', d_details['發言人電話']],
        ['主旨', d_details['主旨']],
        ['符合條款', d_details['符合條款']],
        ['事實發生日', d_details['事實發生日']]]
    for row in data:
        sheet.append(row) #輸入資料
    for i in range(0, len(d_details['說明'])):
        if(i == 0):
            sheet.append(['說明', d_details['說明'][i]])
        else:
            sheet.append(['', d_details['說明'][i]])

def input_data2(Listed_id, Listed_year,d_details,predate,x,k,btn_details):#json,txt
    _filename = str(Listed_id + '_' + Listed_year )
    sheet= {}
    colarray = []
    arraydata={}
    arraydata["序號"]=d_details['序號']
    arraydata["發言日期"]=d_details['發言日期']
    arraydata["發言時間"]=d_details['發言時間']
    arraydata["發言人"]= d_details['發言人']
    arraydata["發言人職稱"]=d_details['發言人職稱']
    arraydata["發言人電話"]=d_details['發言人電話']
    arraydata["主旨"]=d_details['主旨']
    arraydata["符合條款"]=d_details['發言人職稱']
    arraydata["事實發生日"]=d_details['事實發生日']
    arraydata["說明"]=d_details['說明']
    colarray.append(arraydata)
    sheet[_filename]=colarray
    data = json.dumps(sheet,ensure_ascii=False)
    date=str(Listed_id) + '_' + d_details['發言日期'].replace('/', '')
    print(predate)
    print(d_details['發言日期'])
    with open (_filename+'.json','a+',encoding='utf-8') as f:
        if (k == 2):
            f.write('['+str(data)+",")
        elif(k==btn_details):
            f.write(str(data)+']')
        else:
            f.write(str(data)+",")
        f.close()  
    if (predate != (d_details['發言日期'])):
        with open (date+'.txt','a+',encoding='utf-8') as f:
            for i in range(0,len(d_details['說明'])):
                f.write(str(d_details['說明'][i])+"\n")
            f.close()
    else:        
        with open (date+'_'+str(d_details['序號'])+'.txt','a+',encoding='utf-8') as f:
            for i in range(0,len(d_details['說明'])):
                f.write(str(d_details['說明'][i])+"\n")
            f.close()
    predate=d_details['發言日期']
    return predate

def CreateExcel(Listed_id, Listed_year, d_details):
    wb = Workbook() #創建第一個工作表
    frist_sheet = wb.active
    datetime = d_details['發言日期'].replace('/', '.')
    sheet_name = datetime + '-' + d_details['序號'] #工作表名稱
    frist_sheet.title = sheet_name
    input_data(frist_sheet, d_details) #資料輸入資料表
    _filename = str(Listed_id + '-' + Listed_year + '.xlsx')
    wb.save(filename = _filename)
    print('excel_name: ', _filename)
    print('sheet_name: ', sheet_name)
    return _filename

def ReadExcel(d_details, excel_name):
    wb = load_workbook(excel_name) #讀取工作表
    datetime = d_details['發言日期'].replace('/', '.')
    sheet_name = datetime + '-' + d_details['序號']
    sheet = wb.create_sheet(sheet_name)
    input_data(sheet, d_details)
    wb.save(filename = excel_name)
    print('excel_name: ', excel_name)
    print('sheet_name: ', sheet_name)

def DataToExcel(isFrist, Listed_id, Listed_year, d_details, excel_name):
    if(isFrist): #建立工作表
        return CreateExcel(Listed_id, Listed_year, d_details)
    else: #讀取工作表並新增工作表
        ReadExcel(d_details, excel_name)

def get_year_message(Listed):
    for i in Listed:
        input_text(i, '//*[@id="co_id"]') #公司代號或簡稱
        for j in year_range_list:
            input_text(j, '//*[@id="year"]') #年度
            print('id: %s\tyear: %s' % (i, j))
            btn_search = page.locator("xpath=//input[@type='button' and @value=' 查詢 ']") #查詢按鈕
            btn_search.click()
            time.sleep(3) #等待3s
            again = True
            predate=''
            while(again):
                if(WebWaitXpath('//*[@id="t05st01_fm"]/table/tbody/tr[2]/td[3]')): #等待元件讀取
                    again = False
                    isFrist = True
                    excel_name = '1.xlsx'
                    window_before = page #獲取來源網頁資訊
                    btn_details = page.locator('xpath=//*[@id="t05st01_fm"]/table/tbody/tr').all() #詳細資料按鈕
                    print(btn_details)
                    for k in range(2, len(btn_details) + 1): #迭代每則重大消息按鈕
                        print('第' + str(k - 1) + '個按鈕')
                        again_data = True
                        x=1
                        c = 1
                        while(again_data):
                            ChangeToPopUpWindow(k) #改變視窗焦點
                            if(get_data() == False):
                                BackToSourceWindow(window_before)
                            else:
                                print('Get Data OK!')
                                again_data = False
                                d_details = get_data()
                        #=========================

                        
                        if(k == 2): #判斷是否為首個「詳細資料」按鈕
                            # excel_name = DataToExcel(isFrist, i, j, d_details, excel_name)
                            # print(type(d_details))
                            with open('123.log', 'w') as f:
                                f.writelines(json.dumps(d_details)+'\n')
                            isFrist = False
                            
                            
                        else:
                            DataToExcel(isFrist, i, j, d_details, excel_name)
                            
                        predate=input_data2(i,j,d_details,predate,x,k,len(btn_details))
                        print(c)
                        BackToSourceWindow(window_before)
                        print(c+1)
                        time.sleep(2) #等待2s 再次搜尋下一年
                        
                else:
                    if (page.locator('xpath=//*[@id="table01"]/center/h3').count()):
                        print('該 %s 公開發行公司不繼續公開發行！' % i)
                        break
                    else:
                        time.sleep(10) #等待10s
                        page.reload() #刷新網頁


if __name__ == '__main__':
    try:
        driver_open()
        year_range_list, stock_Id_TWSE_Listed, stock_Id_TPEx_Listed = init()
        get_year_message(stock_Id_TWSE_Listed) #上市公司
        #get_year_message(stock_Id_TPEx_Listed) #上櫃公司
    except:
        pass
        driver_close()